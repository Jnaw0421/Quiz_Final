import openpyxl

def read_excel_file(file_path):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    quiz = []
    for row in range(1, sheet.max_row + 1):
        question = sheet.cell(row, 1).value
        options = []
        for col in range(2, 6):
            options.append(sheet.cell(row, col).value)
        correct_option = sheet.cell(row, 6).value
        explaination = sheet.cell(row,7).value
        quiz.append({
            "question": question,
            "options": options,
            "correct_option": correct_option,
            "explaination": explaination,
        })
    return quiz

def display_quiz(quiz):
    quiz_data = []
    for index, q in enumerate(quiz):
        question = {
            "question_number": index + 1,
            "question": q['question'],
            "options": q['options'],
            "correct_option": q['correct_option'],
            "explaination": q['explaination']
        }
        quiz_data.append(question) 
    return quiz_data

def display_score(quiz, answers):
    score = 0
    incorrect_answers = []
    for index, q in enumerate(quiz):
        if answers[index] and answers[index] == q['correct_option']:
            score += 1
        else:
            incorrect_answer = {
                "question": q['question'],
                "incorrect_answer": answers[index],
                "selected_option": answers[index],
                "correct_answer": q['correct_option'],
                "explaination": q['explaination']
            }
            incorrect_answers.append(incorrect_answer)
    return score, incorrect_answers



def run_quiz(file_path):
    quiz = read_excel_file(file_path)
    return display_quiz(quiz)

if __name__ == '__main__':
    run_quiz("questions.xlsx")
